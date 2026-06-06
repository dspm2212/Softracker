"""Reporting helpers for stored monitoring Parquet files.

Reads the accepted files written by the upload endpoint, enriches each row with
the room code inferred from the stored filename, builds compact summaries, and
exports an Excel workbook for offline analysis.

Author: Daniel Perez
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


DETAIL_COLUMNS = [
    "sala_codigo",
    "hostname",
    "usuario",
    "timestamp_utc",
    "nombre_proceso",
    "nombre_ejecutable",
    "titulo_ventana",
    "memoria_mb",
    "pid",
    "ruta_ejecutable",
    "archivo_origen",
]

EXCEL_MAX_ROWS = 1_048_576


@dataclass(frozen=True)
class ReportFilters:
    """Optional filters for report data."""

    fecha_inicio: date | None = None
    fecha_fin: date | None = None
    sala_codigo: str | None = None
    display_tz: str = "UTC"


def _get_tz(tz_name: str):
    """Return a ZoneInfo object, falling back to UTC on invalid names."""
    try:
        return ZoneInfo(tz_name)
    except (ZoneInfoNotFoundError, KeyError):
        return timezone.utc


def _to_local(dt: datetime | None, tz) -> datetime | None:
    """Convert a UTC-aware datetime to the display timezone."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(tz)


def load_monitoring_rows(
    base_dir: Path,
    parquet_subdir: str = "parquet",
    filters: ReportFilters | None = None,
) -> list[dict[str, Any]]:
    """Load accepted Parquet rows from storage and enrich them with room codes."""
    filters = filters or ReportFilters()
    root = base_dir / parquet_subdir
    if not root.exists():
        return []

    rows: list[dict[str, Any]] = []
    tz = _get_tz(filters.display_tz)
    for path in sorted(root.rglob("*.parquet")):
        room_code = _room_code_from_filename(path)
        if filters.sala_codigo and room_code != filters.sala_codigo:
            continue

        try:
            table = pq.read_table(path)
        except Exception:
            continue


        for row in table.to_pylist():
            ts = _coerce_datetime(row.get("timestamp_utc"))
            if not _date_in_range(ts, filters):
                continue

            enriched = {column: row.get(column) for column in DETAIL_COLUMNS}
            enriched["sala_codigo"] = room_code
            enriched["timestamp_utc"] = _to_local(ts, tz) or row.get("timestamp_utc")
            enriched["archivo_origen"] = str(path.relative_to(base_dir))
            rows.append(enriched)

    rows.sort(
        key=lambda r: (
            r.get("sala_codigo") or "",
            r.get("hostname") or "",
            _datetime_sort_key(r.get("timestamp_utc")),
            r.get("nombre_proceso") or "",
        )
    )
    return rows


def build_report(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Build dashboard and Excel-ready summaries from detail rows."""
    room_stats: dict[str, dict[str, Any]] = defaultdict(_new_group_stats)
    machine_stats: dict[tuple[str, str], dict[str, Any]] = defaultdict(_new_group_stats)
    app_stats: dict[str, dict[str, Any]] = defaultdict(_new_app_stats)

    all_rooms: set[str] = set()
    all_machines: set[str] = set()
    all_samples: set[tuple[str, str, str]] = set()
    active_rows = 0
    idle_rows = 0
    timestamps: list[datetime] = []

    for row in rows:
        room = row.get("sala_codigo") or "SIN-SALA"
        host = row.get("hostname") or "SIN-HOST"
        process = row.get("nombre_proceso")
        ts = _coerce_datetime(row.get("timestamp_utc"))
        sample_key = (room, host, ts.isoformat() if ts else "")
        is_active = bool(process)

        all_rooms.add(room)
        all_machines.add(host)
        all_samples.add(sample_key)
        if ts:
            timestamps.append(ts)

        if is_active:
            active_rows += 1
        else:
            idle_rows += 1

        _add_to_group(room_stats[room], row, sample_key, is_active)
        _add_to_group(machine_stats[(room, host)], row, sample_key, is_active)

        if is_active:
            app = str(process)
            app_stats[app]["registros"] += 1
            app_stats[app]["salas"].add(room)
            app_stats[app]["equipos"].add(host)
            memoria = _safe_float(row.get("memoria_mb"))
            if memoria is not None:
                app_stats[app]["memoria_total"] += memoria
                app_stats[app]["memoria_count"] += 1

    room_rows = [_finalize_group(room, stats) for room, stats in room_stats.items()]
    machine_rows = [
        _finalize_group(f"{room} / {host}", stats, room=room, hostname=host)
        for (room, host), stats in machine_stats.items()
    ]
    app_rows = [_finalize_app(name, stats) for name, stats in app_stats.items()]

    room_rows.sort(key=lambda r: r["sala_codigo"])
    machine_rows.sort(key=lambda r: (r["sala_codigo"], r["hostname"]))
    app_rows.sort(key=lambda r: (-r["registros"], r["nombre_proceso"]))

    summary = {
        "total_registros": len(rows),
        "salas": len(all_rooms),
        "equipos": len(all_machines),
        "capturas": len(all_samples),
        "registros_apps": active_rows,
        "capturas_sin_apps": idle_rows,
        "primera_captura": _format_dt(min(timestamps) if timestamps else None),
        "ultima_captura": _format_dt(max(timestamps) if timestamps else None),
    }

    return {
        "summary": summary,
        "rooms": room_rows,
        "machines": machine_rows,
        "apps": app_rows,
        "rows": rows,
    }


def export_report_excel(report: dict[str, Any]) -> bytes:
    """Create an Excel workbook with summaries, detail, and one sheet per room."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary_sheet(workbook, report)
    _write_table_sheet(workbook, "Salas", report["rooms"])
    _write_table_sheet(workbook, "Aplicaciones", report["apps"])
    _write_table_sheet(workbook, "Equipos", report["machines"])
    _write_detail_sheet(workbook, "Datos", report["rows"])

    used_sheet_names = set(workbook.sheetnames)
    by_room: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in report["rows"]:
        by_room[row.get("sala_codigo") or "SIN-SALA"].append(row)

    for room, rows in sorted(by_room.items()):
        sheet_name = _unique_sheet_name(_safe_sheet_name(room), used_sheet_names)
        _write_detail_sheet(workbook, sheet_name, rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_summary_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Resumen")
    sheet["A1"] = "Resumen general"
    sheet["A1"].font = Font(bold=True, size=16, color="1F2937")
    sheet["A3"] = "Indicador"
    sheet["B3"] = "Valor"
    _style_header_row(sheet, 3, 2)

    labels = [
        ("Salas", report["summary"]["salas"]),
        ("Equipos", report["summary"]["equipos"]),
        ("Capturas", report["summary"]["capturas"]),
        ("Registros de aplicaciones", report["summary"]["registros_apps"]),
        ("Capturas sin aplicaciones", report["summary"]["capturas_sin_apps"]),
        ("Total registros", report["summary"]["total_registros"]),
        ("Primera captura", report["summary"]["primera_captura"] or ""),
        ("Ultima captura", report["summary"]["ultima_captura"] or ""),
    ]
    for item in labels:
        sheet.append(item)

    sheet["D3"] = "Top aplicaciones"
    sheet["E3"] = "Registros"
    _style_header_row(sheet, 3, 5, start_col=4)
    for idx, app in enumerate(report["apps"][:10], start=4):
        sheet.cell(row=idx, column=4, value=app["nombre_proceso"])
        sheet.cell(row=idx, column=5, value=app["registros"])

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 14
    sheet.freeze_panes = "A4"


def _write_table_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["Sin datos"])
        sheet.column_dimensions["A"].width = 24
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows[: EXCEL_MAX_ROWS - 1]:
        sheet.append([_excel_value(row.get(header)) for header in headers])

    _finish_table_sheet(sheet, title.replace("-", "_").replace(" ", "_"))


def _write_detail_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(DETAIL_COLUMNS)
    max_detail_rows = EXCEL_MAX_ROWS - 1
    for row in rows[:max_detail_rows]:
        sheet.append([_excel_value(row.get(column)) for column in DETAIL_COLUMNS])

    if len(rows) > max_detail_rows:
        sheet.append(["TRUNCADO", f"Se omitieron {len(rows) - max_detail_rows} filas"])

    _finish_table_sheet(sheet, title.replace("-", "_").replace(" ", "_"))


def _finish_table_sheet(sheet, table_name: str) -> None:
    max_row = max(sheet.max_row, 1)
    max_col = max(sheet.max_column, 1)
    _style_header_row(sheet, 1, max_col)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if max_row > 1:
        ref = f"A1:{sheet.cell(row=max_row, column=max_col).coordinate}"
        table = Table(displayName=_safe_table_name(table_name), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    widths = {
        "A": 18,
        "B": 20,
        "C": 18,
        "D": 22,
        "E": 20,
        "F": 22,
        "G": 34,
        "H": 14,
        "I": 12,
        "J": 46,
        "K": 42,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def _style_header_row(sheet, row_idx: int, max_col: int, start_col: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    for col in range(start_col, max_col + 1):
        cell = sheet.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def _room_code_from_filename(path: Path) -> str:
    return path.name.split("_", 1)[0] if "_" in path.name else "SIN-SALA"


def _date_in_range(ts: datetime | None, filters: ReportFilters) -> bool:
    if ts is None:
        return filters.fecha_inicio is None and filters.fecha_fin is None
    value = ts.date()
    if filters.fecha_inicio and value < filters.fecha_inicio:
        return False
    if filters.fecha_fin and value > filters.fecha_fin:
        return False
    return True


def _coerce_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None


def _datetime_sort_key(value: Any) -> str:
    dt = _coerce_datetime(value)
    return dt.isoformat() if dt else ""


def _format_dt(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _safe_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _new_group_stats() -> dict[str, Any]:
    return {
        "registros": 0,
        "registros_apps": 0,
        "capturas_sin_apps": 0,
        "capturas": set(),
        "equipos": set(),
        "apps": Counter(),
        "timestamps": [],
    }


def _new_app_stats() -> dict[str, Any]:
    return {
        "registros": 0,
        "salas": set(),
        "equipos": set(),
        "memoria_total": 0.0,
        "memoria_count": 0,
    }


def _add_to_group(
    stats: dict[str, Any],
    row: dict[str, Any],
    sample_key: tuple[str, str, str],
    is_active: bool,
) -> None:
    stats["registros"] += 1
    stats["capturas"].add(sample_key)
    stats["equipos"].add(row.get("hostname") or "SIN-HOST")
    ts = _coerce_datetime(row.get("timestamp_utc"))
    if ts:
        stats["timestamps"].append(ts)
    if is_active:
        stats["registros_apps"] += 1
        stats["apps"][row.get("nombre_proceso")] += 1
    else:
        stats["capturas_sin_apps"] += 1


def _finalize_group(
    name: str,
    stats: dict[str, Any],
    room: str | None = None,
    hostname: str | None = None,
) -> dict[str, Any]:
    top_app = stats["apps"].most_common(1)
    result: dict[str, Any] = {
        "sala_codigo": room or name,
        "registros": stats["registros"],
        "capturas": len(stats["capturas"]),
        "equipos": len(stats["equipos"]),
        "registros_apps": stats["registros_apps"],
        "capturas_sin_apps": stats["capturas_sin_apps"],
        "app_mas_frecuente": top_app[0][0] if top_app else "",
        "primera_captura": _format_dt(min(stats["timestamps"]) if stats["timestamps"] else None),
        "ultima_captura": _format_dt(max(stats["timestamps"]) if stats["timestamps"] else None),
    }
    if hostname is not None:
        result = {"hostname": hostname, **result}
    return result


def _finalize_app(name: str, stats: dict[str, Any]) -> dict[str, Any]:
    count = stats["memoria_count"]
    return {
        "nombre_proceso": name,
        "registros": stats["registros"],
        "salas": len(stats["salas"]),
        "equipos": len(stats["equipos"]),
        "memoria_mb_promedio": round(stats["memoria_total"] / count, 2) if count else "",
    }


def _excel_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("-" if char in "[]:*?/\\\"" else char for char in name).strip()
    return (cleaned or "Hoja")[:31]


def _unique_sheet_name(name: str, used: set[str]) -> str:
    candidate = name[:31]
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{name[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def _safe_table_name(name: str) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in name)
    cleaned = cleaned.strip("_") or "Tabla"
    if cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:200]
